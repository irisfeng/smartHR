from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models import Candidate
import tempfile
import os

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "template.xlsx")

COLUMN_MAP = [
    ("sequence_no", "序号"),
    ("recommend_date", "推荐日期"),
    ("recommend_channel", "推荐渠道"),
    ("name", "姓名"),
    ("id_number", "身份证"),
    ("age", "年龄"),
    ("gender", "性别"),
    ("phone", "电话"),
    ("education", "学历"),
    ("school", "毕业学校"),
    ("major", "专业"),
    ("screening_date", "筛选日期"),
    ("leader_screening", "领导初筛"),
    ("screening_result", "筛选邀约结果"),
    ("interview_date", "面试日期"),
    ("interview_time", "面试时间"),
    ("interview_note", "备注"),
    ("first_interview_result", "一面结果"),
    ("first_interview_note", "备注"),
    ("second_interview_invite", "二面邀约"),
    ("second_interview_result", "二面结果"),
    ("second_interview_note", "备注"),
    ("project_transfer", "转项目"),
]

def generate_excel(position_id: int, db: Session) -> str:
    candidates = (
        db.query(Candidate)
        .filter(Candidate.job_position_id == position_id)
        .order_by(Candidate.sequence_no)
        .all()
    )
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for col_idx, (_, header) in enumerate(COLUMN_MAP, 1):
            ws.cell(row=1, column=col_idx, value=header)
    for row_idx, candidate in enumerate(candidates, 2):
        for col_idx, (field, _) in enumerate(COLUMN_MAP, 1):
            value = getattr(candidate, field, "")
            ws.cell(row=row_idx, column=col_idx, value=value if value else "")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
