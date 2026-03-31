import os
import pytest
from openpyxl import load_workbook

from app.models import JobPosition, Candidate
from app.services.export_service import generate_excel, COLUMN_MAP


EXPECTED_HEADERS = [header for _, header in COLUMN_MAP]


def _make_position(db, user) -> JobPosition:
    """Helper: create a minimal JobPosition and return it."""
    pos = JobPosition(
        title="Test Engineer",
        department="Engineering",
        description="Test position",
        created_by=user.id,
    )
    db.add(pos)
    db.commit()
    db.refresh(pos)
    return pos


def _make_candidate(db, position_id: int, seq: int, name: str) -> Candidate:
    """Helper: create a Candidate with a few populated fields."""
    c = Candidate(
        job_position_id=position_id,
        resume_file_path="/fake/resume.pdf",
        sequence_no=seq,
        name=name,
        gender="男",
        phone="13800000000",
        education="本科",
        school="Test University",
        major="Computer Science",
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


# ---------- Unit tests for generate_excel ----------


def test_generate_excel_no_template(db, hr_user, monkeypatch):
    """When template.xlsx does not exist, a workbook is created from scratch
    with the correct 23 header columns."""
    monkeypatch.setattr(
        "app.services.export_service.TEMPLATE_PATH",
        "/nonexistent/path/template.xlsx",
    )
    pos = _make_position(db, hr_user)
    path = generate_excel(pos.id, db)

    try:
        assert os.path.isfile(path)
        wb = load_workbook(path)
        ws = wb.active

        # Verify all 23 headers present in row 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMN_MAP) + 1)]
        assert headers == EXPECTED_HEADERS
        assert len(headers) == 23

        # No data rows (position has no candidates)
        assert ws.cell(row=2, column=1).value is None
    finally:
        os.unlink(path)


def test_generate_excel_with_candidates(db, hr_user, monkeypatch):
    """Candidates are written into the workbook in sequence_no order and
    the file is valid XLSX readable by openpyxl."""
    monkeypatch.setattr(
        "app.services.export_service.TEMPLATE_PATH",
        "/nonexistent/path/template.xlsx",
    )
    pos = _make_position(db, hr_user)
    c2 = _make_candidate(db, pos.id, seq=2, name="Bob")
    c1 = _make_candidate(db, pos.id, seq=1, name="Alice")

    path = generate_excel(pos.id, db)

    try:
        wb = load_workbook(path)
        ws = wb.active

        # Row 2 should be Alice (sequence_no=1), Row 3 should be Bob (sequence_no=2)
        assert ws.cell(row=2, column=1).value == 1  # sequence_no
        assert ws.cell(row=2, column=4).value == "Alice"  # name
        assert ws.cell(row=3, column=1).value == 2
        assert ws.cell(row=3, column=4).value == "Bob"

        # Verify a few more mapped fields for the first candidate
        assert ws.cell(row=2, column=7).value == "男"      # gender
        assert ws.cell(row=2, column=8).value == "13800000000"  # phone
        assert ws.cell(row=2, column=9).value == "本科"    # education
        assert ws.cell(row=2, column=10).value == "Test University"  # school
        assert ws.cell(row=2, column=11).value == "Computer Science"  # major

        # Row 4 should be empty (only 2 candidates)
        assert ws.cell(row=4, column=1).value is None
    finally:
        os.unlink(path)


# ---------- Integration tests via the HTTP endpoint ----------


def test_export_endpoint(client, db, hr_user, hr_headers, monkeypatch):
    """GET /api/positions/{id}/export returns 200 with the correct XLSX
    content type."""
    monkeypatch.setattr(
        "app.services.export_service.TEMPLATE_PATH",
        "/nonexistent/path/template.xlsx",
    )
    pos = _make_position(db, hr_user)
    _make_candidate(db, pos.id, seq=1, name="Charlie")

    resp = client.get(f"/api/positions/{pos.id}/export", headers=hr_headers)

    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Response body should start with the XLSX magic bytes (PK zip header)
    assert resp.content[:2] == b"PK"


def test_export_empty_position(client, db, hr_user, hr_headers, monkeypatch):
    """Exporting a position with zero candidates still returns a valid XLSX
    file containing only the header row."""
    monkeypatch.setattr(
        "app.services.export_service.TEMPLATE_PATH",
        "/nonexistent/path/template.xlsx",
    )
    pos = _make_position(db, hr_user)

    resp = client.get(f"/api/positions/{pos.id}/export", headers=hr_headers)

    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"

    # Write to a temp file so openpyxl can verify the content
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        tmp.write(resp.content)
        tmp.close()
        wb = load_workbook(tmp.name)
        ws = wb.active

        # Headers are present
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMN_MAP) + 1)]
        assert headers == EXPECTED_HEADERS

        # No data rows
        assert ws.cell(row=2, column=1).value is None
    finally:
        os.unlink(tmp.name)


def test_export_as_manager_forbidden(client, db, manager_user, manager_headers, monkeypatch):
    """Managers cannot export — only HR can (contains sensitive PII)."""
    monkeypatch.setattr(
        "app.services.export_service.TEMPLATE_PATH",
        "/nonexistent/path/template.xlsx",
    )
    pos = _make_position(db, manager_user)
    resp = client.get(f"/api/positions/{pos.id}/export", headers=manager_headers)
    assert resp.status_code == 403
